import os
import win32com.client as win32
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.chart import BarChart, PieChart, Reference
import shutil

# Create Outlook app instance
outlook=win32.Dispatch("Outlook.Application").GetNamespace("MAPI")
inbox=outlook.Folders("niko_ninja@outlook.com").Folders('inbox') #load inbox and save to variable

messages=inbox.Items
save_folder=r"F:\Nikhil\study\tech\Programming\python\projects\yt-py-mail-excel-filefolder\downloaded excels"

# Clear old files if the download folder exists
if os.path.exists(save_folder):
    for f in os.listdir(save_folder):
        os.remove(os.path.join(save_folder, f))
else:
    os.makedirs(save_folder)


for msg in messages:
    if 'sales data' in str(msg.Subject): #filter by subject
        print(msg.Subject)

    # filter by partial msg on subject, email address
    if 'sales data' in msg.Subject and msg.SenderEmailAddress=='niko_ninja@outlook.com':
        print(msg.Subject+"  " + str(msg.ReceivedTime))

        # save attachment

        if not os.path.exists(save_folder):  # check if folder exists and create path
            os.makedirs(save_folder)  #if doesnt exist, create folder
        for atch in msg.Attachments:
            # Split "image.png" into ("image", ".png")
            base_name, extension = os.path.splitext(atch.FileName)

            # Standard initial path
            file_path = os.path.join(save_folder, atch.FileName)

            # Counter for duplicates
            counter = 1

            while os.path.exists(file_path):
                new_filename = f"{base_name}_{counter}{extension}"
                file_path = os.path.join(save_folder, new_filename)
                counter += 1

            # Save using the unique path
            atch.SaveAsFile(file_path)

#-----------------------------
src_folder=r"F:\Nikhil\study\tech\Programming\python\projects\yt-py-mail-excel-filefolder\downloaded excels"
dest_folder=r"F:\Nikhil\study\tech\Programming\python\projects\yt-py-mail-excel-filefolder\master excels"

#check if destnation folder exists
os.makedirs(dest_folder, exist_ok=True)

# Generate Master File with timestamp
timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
file_name = f"Masterfile_{timestamp}.xlsx"
master_file_path=os.path.join(dest_folder, file_name)

# Create a master workbook
wb = Workbook()
ws = wb.active
ws.title = "Full report"

# Add header and content
ws.append(["Location", "Revenue","region"])

# counter for region - dashboard
regions_found = []

for file in os.listdir(src_folder):
    # Only process Excel files and skip temp hidden files (~$)
    if file.endswith(".xlsx") and not file.startswith("~$"):

        region_name = os.path.splitext(file)[0] #extract region name
        regions_found.append(region_name)

        src_file_path = os.path.join(src_folder, file) #create full name of file
        source_wb = load_workbook(src_file_path, data_only=True)
        source_ws = source_wb.active

    # Read data starting from row 2 (skipping header)
        for row in source_ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                row_list = list(row)

            # Add region_name in 3rd column
                row_list.append(region_name)

            # Append row to Master sheet
                ws.append(row_list)

        source_wb.close()
#-----------------------------
# create dashboard for pivot and charts
ws_dashboard = wb.create_sheet(title="Dashboard")
ws_dashboard.append(["Region", "Total Revenue"])

# SUMIF formulas for each region
row_num = 2
for region in sorted(regions_found):
    #  SUMIF formula at 'Full report' Column C (Region) and Column B (Revenue)
    formula = f"=SUMIF('Full report'!C:C, A{row_num}, 'Full report'!B:B)"
    ws_dashboard.append([region, formula])
    row_num += 1

max_row = row_num - 1

#Add bar chart
barchart = BarChart()
barchart.type = "col"
barchart.style = 10
barchart.title = "Revenue by Region"
barchart.y_axis.title = "Revenue"
barchart.x_axis.title = "Region"

# Reference data from Dashboard sheet
chart_data = Reference(ws_dashboard, min_col=2, min_row=1, max_row=max_row)
chart_labels = Reference(ws_dashboard, min_col=1, min_row=2, max_row=max_row)

barchart.add_data(chart_data, titles_from_data=True)
barchart.set_categories(chart_labels)
ws_dashboard.add_chart(barchart, "D2")

 #Add bar chart

piechart = PieChart()
piechart.title = "Regional Revenue Share"

piechart.add_data(chart_data, titles_from_data=True)
piechart.set_categories(chart_labels)
ws_dashboard.add_chart(piechart, "D18")


# 4. Save updated Master file
wb.save(master_file_path)
print(f"Successfully combined files into {master_file_path}!")

#-----------------------------

import os
import win32com.client as win32

# Create Outlook app instance
olApp = win32.Dispatch('outlook.Application')
olNS = olApp.GetNameSpace('MAPI')  # Fixed spelling: GetNameSpace

# Construct the email item object
mailItem = olApp.CreateItem(0)
mailItem.Subject = 'Quarterly master report'
mailItem.BodyFormat = 1  # 1 = Plain Text
mailItem.Body = ("Hello team,"
                "\n"
                 "\nplease find attached Master report for this Quarter"
                 "\n"
                 "\n"
                 "\n"
                 "\nThanks,"
                 "\nNiko")
mailItem.To = 'niko.ninja@outlook.com'

# Attach files located in the current directory
mailItem.Attachments.Add(master_file_path)

# Display, Save to Drafts, and Send
mailItem.Display()
mailItem.Save()